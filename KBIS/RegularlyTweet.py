import  twitter
import  datetime
import  openpyxl
import  LogWriterClassVer
import traceback
class Tweets(object):
    def __init__(self,api:twitter.Api,inDataBaseMoneyListsPath:str,logpath:str):
        self.logwriter=LogWriterClassVer.LogWriterClassVer(logpath)
        self.api=api
        self.path=inDataBaseMoneyListsPath
        self.todayIsSent=False
        self.ResetHour=11
        self.RegularlyTweetHour=12
        self.INCOMESTART_row=7#(roop)
        self.INCOME_column=7#G
        self.OUTGOSTART_row=7#(roop)
        self.OUTGO_column=9#I
        #
        self.INCOME_END=999
        self.OUTGO_END=999
    def Check(self):
        if((datetime.datetime.now().hour==self.RegularlyTweetHour) and (not self.todayIsSent)):
            try:
                self.api.PostUpdate(f'{datetime.datetime.now()}\n【定期更新】\n現在の会計残高は{str(self.MakeNowFundTweet())}円です。')
                self.logwriter.LogWrite('print','定期ツイートしました。')
                self.todayIsSent = True
            except:
                self.logwriter.LogWrite('print',f'定期更新に失敗 traceback:\n{traceback.format_exc()}')
        elif((datetime.datetime.now().hour==self.ResetHour) and (self.todayIsSent)):
            self.todayIsSent=False
    def MakeNowFundTweet(self)->int:
        wb=openpyxl.load_workbook(self.path)
        sheet=wb['出納']
        INCOME_sum=0
        for now_row in range(self.INCOMESTART_row,self.INCOME_END):
            try:
                INCOME_sum+=sheet.cell(row=now_row,column=self.INCOME_column).value
            except:
                pass
        print(f'income(sum)={INCOME_sum}')
        OUTGO_sum=0
        for now_row in range(self.OUTGOSTART_row,self.OUTGO_END):
            try:
                OUTGO_sum+=sheet.cell(row=now_row,column=self.OUTGO_column).value
            except:
                pass
        print(f'OUTGO(sum)={OUTGO_sum}')
        print(f'残高={INCOME_sum-OUTGO_sum}')
        return INCOME_sum-OUTGO_sum