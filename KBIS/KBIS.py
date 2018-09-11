# coding:utf-8
from __future__ import unicode_literals
#フォローは手作業でお願いしますね・・・
import twitter
import __future__
import os
import sys
import openpyxl
import oauth2
import webbrowser
import time
import datetime
import RootCommand
import traceback
import LogWriterClassVer
import APIKeyReader
l=LogWriterClassVer.LogWriterClassVer()
dir="DirectryLineEmpty"#初期化用
BOOKNAME="個人支払出納管理簿.xlsx"#読み込むxlsxファイル
#BOOKNAME="個人支払管理簿_ModifiedExample.xlsx"
TWITTER_BE_COMPATIBLE_BY_THIS_BOOK="Twitter対応リスト.xlsx"#読み込むxlsxファイル
ONE_SHEET_MAX_MEMBER=20#一シート当たりのメンバー数の最大(for文で使う)
MAX_USER=1000#最大利用者数
MAX_SHUTUNOU=100#出納数の最大
IGNOREFILE="IgnoreList.txt"#IgnoreListのファイル名

def RootCommandPlayer(user,api,raw_str):#raw_strにはcommand,argumentが入ってる。
    if(not RootChallenger):
        return
def CommandReader(command):
    if(not RootChallenger):
        return
    _command=command.split(" ")
    print(str(_command)+"and size"+str(len(_command)))
    if(len(_command)<2):
        api.PostDirectMessage(screen_name=api.VerifyCredentials().screen_name,text="rootコマンドとして多分おかしいんじゃないかな？")
    elif(str(len(_command))=="2"):
         rootcomand.Renew(_command.pop(0),_command.pop(0))
    elif(str(len(_command))=="3"):
         rootcomand.Renew(_command.pop(0),_command.pop(0),_command.pop(0))
    elif(str(len(_command))=="4"):
         rootcomand.Renew(_command.pop(0),_command.pop(0),_command.pop(0),_command.pop(0))
    elif(str(len(_command))=="5"):
         rootcomand.Renew(_command.pop(0),_command.pop(0),_command.pop(0),_command.pop(0),_command.pop(0))
    elif(str(len(_command))=="6"):
         rootcomand.Renew(_command.pop(0),_command.pop(0),_command.pop(0),_command.pop(0),_command.pop(0),_command.pop(0))     
    else:
        api.PostDirectMessage(screen_name=api.VerifyCredentials().screen_name,text="引数が多すぎる")
def RootChallenger(user,api):
    #rootコマンドを頼んだuserが正当か判断する。
    #Twitter対応リストのIDの横のセルに「su」と入力すること
    #7/1ここデバッグしてるので汚い
    TWIT_BOOK_SHEET=GetTwitterSheet()
    exsist=False
    for _row in range(2,MAX_USER,1):
        print("比較中...{0} and {1}".format(TWIT_BOOK_SHEET.cell(row=_row,column=2).value,user))
        if(str(TWIT_BOOK_SHEET.cell(row=_row,column=2).value)==user):
            print("エクセルファイルにユーザーがいました。！")
            exsist=True
            if(str(TWIT_BOOK_SHEET.cell(row=_row,column=3).value)=="su"):
                l.LogWrite("print","SuperUser認証が通りました！")
                return True
    if(exsist):
        l.LogWrite("FalseSU",user)
    l.LogWrite("print","SuperUser要求が{0}から来ました。そのユーザーはエクセルファイルにいたか?:{1}。SuperUserではありませんでした。".format(user,exsist))
    api.PostDirectMessage(screen_name=api.VerifyCredentials().screen_name, text="SuperUser要求が来たけどSuperUserじゃなかったです。要求者:{0}".format(user))
    return False
def CheckDirFile():
    print(os.getcwd()+"にある、SetDirectry.txtファイルを読み込みます");
    try:
        DIRECTRY=open("SetDirectry.txt")
    except:
        print("ファイルのオープンに失敗しました")
        print("新規ファイルを作成しますか？(Y/N)")
        ans=input()
        while((ans!="y" and ans!="n")and(ans!="Y")and(ans!="N")):
            print("再度入力してください。")
            print("新規ファイルを作成しますか？(Y/N)")
            ans=input()
        if(ans=="N"or ans=="n"):
            print("プログラムを終了します。Push Any Key")
            ans=input()
            sys.exit()
        else:
            print("SetDirectryファイルを作成します")
            DIRECTRY=open("SetDirectry.txt",'w')
            DIRECTRY.write(os.getcwd())
            print(os.getcwd()+"とSetDirectry.txtに書き込み、保存しました。")
            print("作業用フォルダを変えたい場合は"+os.getcwd()+"のSetDirectry.txtファイルの内容を変更してください")
            print("プログラムを終了します。Push Any Key ")
            DIRECTRY.close()
            ans=input()
            sys.exit()
    dir=DIRECTRY.readline()
    try:
        os.chdir(dir)
    except:
        print("SetDirectry.txtの内容が不正です。Push Any Key")
        ans=input()
        traceback.print_exc()
        sys.exit()
    print("読み込み完了.作業用ディレクトリは「"+dir+"」です。")
    return dir
class MemberData(object):
    def __init__(self,name,money):
        self.name=name
        self.money=money
        self.twiiterID="none"
        self.remarks="none"
    def outputPrototype(self):
        print(self.name+","+str(self.money)+","+self.twiiterID+self.remarks)
def MessageParser(memberList,memberName):#メンバーリストからメンバーを見つけた後、返信する文章を作成する関数
    for t in range(len(memberList)):
        if(memberList[t].name==memberName):
            if(memberList[t].money==0):
                statusStr="未納額はありません。"
            elif(memberList[t].money<0):
                statusStr=str((-1)*memberList[t].money)+"円を返金します。"
            else:
                statusStr=str(memberList[t].money)+"円が未納です。"
            statusStr+="\r\n"
            if(memberList[t].remarks!="none"):
                statusStr+="備考:"+str(memberList[t].remarks)
            else:
                statusStr+="備考:なし"
            statusStr+="\r\n"
            statusStr+="\r\n"
            statusStr+="納金・返金は定例会または都合のいい活動日に対応いたします。ご質問ご相談などあれば会計まで。"
            print(statusStr)
            return statusStr
def GetSheetList():
    try:
        DATA_BOOK=openpyxl.load_workbook(BOOKNAME)
    except:
        print(BOOKNAME+"を開くのに失敗しました。　Push Any Key")
        ans=input()
        sys.exit()
    sheetList=[]
    #年度更新
    sheet15G=DATA_BOOK["15G"]
    sheet16G=DATA_BOOK["16G"]
    sheet17G=DATA_BOOK["17G"]
    sheet18G=DATA_BOOK["18G"]
    sheet19G=DATA_BOOK["19G"]
    sheet20G=DATA_BOOK["20G"]
    sheetList.append(sheet15G)
    sheetList.append(sheet16G)
    sheetList.append(sheet17G)
    sheetList.append(sheet18G)
    sheetList.append(sheet19G)
    sheetList.append(sheet20G)
    print("sheetListの読み込みが完了")
    return sheetList
def GetTwitterSheet():
    try:
        TWITTER_BOOK=openpyxl.load_workbook(TWITTER_BE_COMPATIBLE_BY_THIS_BOOK)#ツイッターアカウントとメンバーの対応表を開く
    except:
        print(BOOKNAME+"を開くのに失敗しました。　Push Any Key")
        ans=input()
        sys.exit()
    TWITTER_BOOK_SHEET=TWITTER_BOOK["Sheet1"]
    return TWITTER_BOOK_SHEET
def GetMember(sheetList):
    nameList=[]
    memberList=[]
    for GenCount in range(0,6,1):
        sheetTemp=sheetList.pop()
        for i in range(4,20,1):
            tempName=sheetTemp.cell(row=i,column=2).value
            if (tempName):
                moneyTemp=0
                nameList.append(tempName)
                for col in range(8,99,1):#STATUS計算
                    if(sheetTemp.cell(row=i,column=col).value):
                        moneyTemp+=sheetTemp.cell(row=i,column=col).value
                tempClass=MemberData(tempName,moneyTemp)
                if(sheetTemp.cell(row=i,column=(5)).value):#remarksの段
                    tempClass.remarks=sheetTemp.cell(row=i,column=(5)).value
                memberList.append(tempClass)
    return nameList, memberList
def Check(memberList,nameList):
    try:
        TWITTER_BOOK=openpyxl.load_workbook(TWITTER_BE_COMPATIBLE_BY_THIS_BOOK)#ツイッターアカウントとメンバーの対応表を開く
    except:
        print(TWITTER_BOOK+"を開くのに失敗しました。　Push Any Key")
        ans=input()
        sys.exit()
    TWITTER_BOOK_SHEET=TWITTER_BOOK["Sheet1"]
    for i in range(len(memberList)):
        memberTemp=nameList[i]
        print(memberTemp)
        isThere=False
        for _row in range(2,MAX_USER,1):#ここから総当たりで検索する。
            if(TWITTER_BOOK_SHEET.cell(row=_row,column=1).value and isThere==False):
                if(TWITTER_BOOK_SHEET.cell(row=_row,column=1).value==memberTemp):#対応表に名前があった場合の処理
                    isThere=True
                    if(TWITTER_BOOK_SHEET.cell(row=_row,column=2).value):#もしTwitterIDも対応表にあったなら
                        for t in range(len(memberList)):#memberListから対応するメンバーを見つける
                            if(memberList[t].name==memberTemp):#見つかった場合の処理
                                memberList[t].twiiterID=TWITTER_BOOK_SHEET.cell(row=_row,column=2).value#twitterIDを代入
            elif(TWITTER_BOOK_SHEET.cell(row=_row,column=1).value or isThere==True):
                memberTemp=nameList[i]#なくてもいいけど、なんか入れとかないとVSがエラー吐くので、結果の変わらないダミーを入れた
            else:
                TWITTER_BOOK_SHEET.cell(row=_row,column=1).value=memberTemp
                isThere=True
    TWITTER_BOOK.save(TWITTER_BE_COMPATIBLE_BY_THIS_BOOK)
def DirectMailCheck(directMail,ignoreList,memberList,api):
    for i in range(len(directMail)):
        ignore=False
        isSent=False
        try:
            TWITTER_BOOK=openpyxl.load_workbook(TWITTER_BE_COMPATIBLE_BY_THIS_BOOK)#ツイッターアカウントとメンバーの対応表を開く
        except:
            print(TWITTER_BOOK+"を開くのに失敗しました。")
            return ignoreList
        TWITTER_BOOK_SHEET=TWITTER_BOOK["Sheet1"]
        for t in range(len(ignoreList)):
            if(int(directMail[i].id)==int(ignoreList[t])):
                print("passed")
                ignore=True
        if(not ignore):
            print("passed failed")
        if(directMail[i].text.find("root:")==0):
            if(not ignore):
                if(RootChallenger(directMail[i].sender_screen_name,api)):
                    print("root Command")
                    ignore=True
                    isSent=True
                    command=directMail[i].text.replace("root:","")
                    CommandReader(command)
                else:
                    api.PostDirectMessage(screen_name=directMail[i].sender_screen_name,text="すいません！そのコマンドは許可されていません！")
                    isSent=True
                    ignore=True
                ignoreList.append(directMail[i].id)
        elif(directMail[i].text.find("info:")!=0):
            if(not ignore):
                print("形式が違うDMが届いています。")
                print("送信者:"+directMail[i].sender_screen_name)
                print("内容:"+directMail[i].text)
                ignoreList.append(directMail[i].id)
                ignore=True
        else:
            if(not ignore):
                print("情報要求を感知")
                print("要求ユーザネーム:"+directMail[i].sender_screen_name+" 名前:"+directMail[i].text[5:])
                tempName=directMail[i].text[5:]
                for col in range(2,MAX_USER,1):
                    if(TWITTER_BOOK_SHEET.cell(row=col,column=1).value==tempName):
                        isSent=True
                        if(TWITTER_BOOK_SHEET.cell(row=col,column=2).value==directMail[i].sender_screen_name or not TWITTER_BOOK_SHEET.cell(row=col,column=2).value):
                            print("正しい要求でした。"+tempName+"へ返信します。")
                            TWITTER_BOOK_SHEET.cell(row=col,column=2).value=directMail[i].sender_screen_name
                            if(rootcomand.GetResponcePlace):
                                api.PostDirectMessage(screen_name=directMail[i].sender_screen_name,text=MessageParser(memberList,TWITTER_BOOK_SHEET.cell(row=col,column=1).value))     
                            ignoreList.append(directMail[i].id)
                            l.LogWrite("send",str("True"),str(directMail[i].sender_screen_name),"成功")
                        else:
                           print("多重アカウント検知")
                           print("なりすまし不正疑惑のあるアカウント:"+str(directMail[i].sender_screen_name))
                           print("被害可能性のあるアカウント:"+TWITTER_BOOK_SHEET.cell(row=col,column=2).value)
                           ignoreList.append(directMail[i].id)
                           if(rootcomand.GetResponcePlace):
                                api.PostDirectMessage(screen_name=directMail[i].sender_screen_name,text="成りすまし防止のため金額照会を中止します。")
                           else:
                                pass
                                api.PostUpdate("成りすまし防止のため金額照会を中止します。")
                           l.LogWrite("send",str("False"),str(directMail[i].sender_screen_name),"なりすまし")
        if(not isSent and not ignore):
            print("エクセルファイルと照合しましたが名前が一致しませんでした。")
            if(rootcomand.GetResponcePlace):
                api.PostDirectMessage(screen_name=directMail[i].sender_screen_name,text="エクセルファイルと照合しましたが名前が一致しませんでした。\r\n形式に沿っているか確認してみてください。\r\n形式: info:苗字（半角スペース）名前")
                print("dummyResponceUpdated")
            else:
                api.PostUpdate("エクセルファイルと照合しましたが名前が一致しませんでした。\r\n形式に沿っているか確認してみてください。\r\n形式: info:苗字（半角スペース）名前")
                print("postupdate")
            ignoreList.append(directMail[i].id)
            print(str(directMail[i].id)+"を追加")
            l.LogWrite("send",str("False"),str("directMail[i].sender_screen_name"),"Excel不一致")
        TWITTER_BOOK.save(TWITTER_BE_COMPATIBLE_BY_THIS_BOOK)
    return ignoreList
def AutoTweet(toSaidSwitch,api):
    now=datetime.datetime.now()
    sum=0
    DATA_BOOK=openpyxl.load_workbook(BOOKNAME)
    if(int('{0:%H}'.format(now))==int(12) and not toDaySaidSwitch):#falseかつhourが12時なら投稿
        print("Test now return this")
        now=datetime.datetime.now()
        SYUTUNOU_SHEET=DATA_BOOK["出納"]
        for i in range(7,MAX_SHUTUNOU,1):
            if(SYUTUNOU_SHEET.cell(row=i,column=7).value):
                sum+=SYUTUNOU_SHEET.cell(row=i,column=7).value
            if(SYUTUNOU_SHEET.cell(row=i,column=9).value):
                sum-=SYUTUNOU_SHEET.cell(row=i,column=9).value
        if(rootcomand.GetAutoTweetState()):
            api.PostUpdate(str(now)+"\r\n"+"【定期更新】"+"\r\n"+"現在の会計残高は"+str("{:,}".format(sum))+"円です。")
        else:
            api.PostDirectMessage(screen_name=api.VerifyCredentials().screen_name,text=str(now)+"\r\n"+"【定期更新】"+"\r\n"+"現在の会計残高は"+str("{:,}".format(sum))+"円です。") 
        l.LogWrite("tweet")
        return True
    elif(int('{0:%H}'.format(now))==int(11)):#hourが11時ならfalseにリセット
        l.LogWrite("reset")
        return False
    elif(not toSaidSwitch):#falseならfalseを返す。（時間外の処理）
        return False
    else:#trueならtrueを返す
        return True
if __name__ == '__main__':
    print("hello!" + str(os.name))
    directry = CheckDirFile()
    apiR = APIKeyReader.Reader(os.getcwd())
    api = apiR.GetApi()
    api.PostDirectMessage("【テスト】KBIS 起動", screen_name=api.VerifyCredentials().screen_name)
    pass
    sheetList = GetSheetList()
    TWITTER_BOOK_SHEET = GetTwitterSheet()
    result = GetMember(sheetList)
    nameList = result[0]
    memberList = result[1]
    print("memberとツイッターアカウントの照合をします。")
    Check(memberList, nameList)  # ここからメンバーのtwitterアカウントとの照合を行う
    # ここでtwitterアカウント対応表との関連付けは終わりとする。
    # ここからtwitter側の処理

    rootcomand = RootCommand.RootCommand(api, memberList, "decoy")
    directMail = api.GetDirectMessages()
    IgnoreListFile = open(IGNOREFILE)
    ignoreList = []
    temp = IgnoreListFile.readline().strip()
    while (temp):
        print("ignoreリストからロード:" + str(temp))
        ignoreList.append(temp)
        temp = IgnoreListFile.readline()
    try:
        IgnoreListFile.close()
    except:
        print("IgnoreListFile.close() was failed")

    ignoreList = DirectMailCheck(directMail, ignoreList, memberList, api)

    # とりあえずここまでで起動時のセットアップを終了です
    # Ignoreファイルの更新
    print("Ignoreファイルの更新を行います。")
    writingData = open(IGNOREFILE, 'w')
    strings = ""
    writingData.write(strings)
    writingData.close()
    writingData = open(IGNOREFILE, 'w')
    for t in range(len(ignoreList) - 1):
        strings += str(ignoreList[t]) + "\n"
    strings += str(ignoreList[len(ignoreList) - 1])
    strings = strings.replace("\n\n", "\n")
    writingData.write(strings)
    writingData.close()
    print("Ignoreファイルの更新は終了しました。")
    # ここで更新終了
    # ここから自動システムになります。
    # ダイレクトメッセージの取得間隔は安定をとって、65秒に1回にしています。
    toDaySaidSwitch = False

    l.LogWrite(type1="login")
    while (True):
        rootcomand.ReList(api, memberList)
        print("定期更新モードです。")
        print("---Ctrl+Dで終了出来ます。---")
        time.sleep(65)
        try:
            toDaySaidSwitch = AutoTweet(toDaySaidSwitch, api)
            l.LogWrite("switch", str("toDaySaidSwitch"))
        except:
            print("定期更新に失敗しました。")
            l.LogWrite("switch", str("toDaySaidSwitch"))
            l.LogWrite("print", traceback.format_exc())
        try:
            directMail = api.GetDirectMessages()
            ignoreList = DirectMailCheck(directMail, ignoreList, memberList, api)
            l.LogWrite("check", str("True"))
        except:
            print("何らかの処理に失敗しています。65秒後に再度更新します。")
            l.LogWrite("print", traceback.format_exc())
            l.LogWrite("check", str("False"))
        # Twitter側の更新はここまで
        # Ignoreファイルの更新はここから
        print("Ignoreファイルの更新を行います。")
        strings = ""
        os.remove(IGNOREFILE)
        writingData = open(IGNOREFILE, 'w')
        for t in range(len(ignoreList) - 1):
            strings += str(ignoreList[t]) + "\n"
        strings += str(ignoreList[len(ignoreList) - 1])
        strings = strings.replace("\n\n", "\n")
        writingData.write(strings)
        writingData.close()
        print("Ignoreファイルの更新は終了しました。行数:" + str(len(ignoreList)))
        # ignoreファイルの更新を終了

