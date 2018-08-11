# encoding :utf-8
import sqlite3
import openpyxl
import traceback
import os
import LINENotifer
class DataBases(object):
    def __init__(self,devmode:bool):
        self.MAXGEN = 22#const
        self.MINGEN = 15
        LINENotifer.Notify.MessageCall('DataBase 1/6 メモリ内のデータベースを確認しています。')
        try:
            os.remove(":memory:")
        except:
            pass
        LINENotifer.Notify.MessageCall('DataBase 2/6 データベースに接続します。')
        if(devmode):
            self.twitterBook = '../Tools/Twitter対応リスト.xlsx'
            self.moneyBook='../Tools/237585_個人支払出納管理簿.xlsx'
        else:
            self.twitterBook = '../../KBIS_Workingplace/Twitter対応リスト.xlsx'
            self.moneyBook='../../KBIS_Workingplace/個人支払出納管理簿.xlsx'
            #self.twitterBook = '../Tools/Twitter対応リスト.xlsx'
            #self.moneyBook='../Tools/237585_個人支払出納管理簿.xlsx'
        self.connect = sqlite3.connect(":memory:")
        LINENotifer.Notify.MessageCall('DataBase 3/6 データベースに接続しました。テーブルの作成、ユーザの追加を行います。')
        self.cursor = self.connect.cursor()
        create_table = '''create table users(gen int,realname TEXT,twittername TEXT,money int,remarks TEXT,authority TEXT,UNIQUE (realname,twittername)) '''
        self.sql = 'insert into users (gen,realname,twittername,money,remarks,authority) values (?,?,?,?,?,?)'
        LINENotifer.Notify.MessageCall('DataBase 4/6 ユーザの追加を終了しました。エクセルファイルを読み込みます。')
        self.cursor.execute(create_table)
        for i in range(self.MINGEN, self.MAXGEN):
            try:
                self.cursor.executemany(self.sql, self.CreateUsersFromSheet(i))
            except KeyError:
                LINENotifer.Notify.MessageCall(traceback.format_exc())
        print('Hello DB')
        LINENotifer.Notify.MessageCall('DataBase 5/6 読み込みが完了しました。')
        select_sql = 'select * from users'
        for row in self.cursor.execute(select_sql):
            print(row)
        LINENotifer.Notify.MessageCall('DataBase 6/6データベースの構築完了')
    def CreateUsersFromSheet(self,gen):  # SQLに追加できるように手に入れたデータを変換する
        userList = []

        moneybook = openpyxl.load_workbook(self.moneyBook)
        sheet=moneybook[f'{gen}G']
        twitterbook = openpyxl.load_workbook(self.twitterBook)
        sheetTB = twitterbook['Sheet1']
        for user in range(1, 100):
            # moneyCreating
            sum = 0
            for debt in range(8, 200):
                if (sheet.cell(row=(user + 3), column=debt).value):
                    try:
                        sum += int(sheet.cell(row=(user + 3), column=debt).value)
                    except TypeError:
                        pass
            # Twitterアカウントとの結びつけ
            exist = False
            authority = None
            for i in range(1, 200):
                if (sheet.cell(row=(user + 3), column=2).value == sheetTB.cell(row=(i+1), column=1).value):
                    twitterName = sheetTB.cell(row=(i+1), column=2).value
                    if (sheetTB.cell(row=(i+1), column=3).value):
                        authority = sheetTB.cell(row=(i+1), column=3).value
                    exist = True
            if(not exist):
                i=0
                LINENotifer.Notify.MessageCall(
                    f'管理簿にいて対応リストにいないUserを発見{sheet.cell(row=(user + 3), column=2).value}')

                while(True):
                    if(sheetTB.cell(row=(i+1), column=1).value):
                        pass
                    else:
                        sheetTB.cell(row=(i+1),column=1).value=sheet.cell(row=(user + 3), column=2).value
                        LINENotifer.Notify.MessageCall(f'管理簿にいて対応リストにいないUserを追加しました。{sheet.cell(row=(user + 3), column=2).value}')
                        twitterName=None
                        authority=None
                        exist=True
                        break
                    i=i+1


            #
            if(sheet.cell(row=(user + 3), column=2).value):
                userList.append((gen, sheet.cell(row=(user + 3), column=2).value, twitterName, sum,sheet.cell(row=(user + 3), column=5).value, authority))
            else:
                pass
        moneybook.save(self.moneyBook)
        twitterbook.save(self.twitterBook)
        return userList
    def renew(self):#一回全部消すか・・・
        delete_usersql='''drop table users'''
        delete_twittersql='''drop table TwitterExistsUser'''
        try:
            self.cursor.execute(delete_usersql)
            self.cursor.execute(delete_twittersql)
        except:
            traceback.print_exc()
        create_table = '''create table if not exists users(gen int,realname TEXT,twittername TEXT,money int,remarks TEXT,authority TEXT,UNIQUE (realname,twittername)) '''
        self.cursor.execute(create_table)
        self.sql = 'insert into users (gen,realname,twittername,money,remarks,authority) values (?,?,?,?,?,?)'
        createTwitterUserTable='''create table if not exists TwitterExistsUser(gen int,realname TEXT,twittername TEXT,money int,remarks TEXT,authority TEXT,UNIQUE (realname,twittername)) '''
        self.cursor.execute(createTwitterUserTable)

        workbook = openpyxl.load_workbook(self.moneyBook)
        for i in range(self.MINGEN, self.MAXGEN):
            try:
                sheet = workbook['{0}G'.format(i)]
                self.cursor.executemany(self.sql, self.CreateUsersFromSheet(sheet, i))
            except KeyError:
                break
        print('ユーザ全体のリストを表示します。')
        for i in self.cursor.execute('''select * from users'''):
            print(i)
        twitterlist=self.Search('at','all')
        print('Twitterユーザのテーブルを更新しています・・・')
        for i in twitterlist:
            print(i)
        print('Twitterユーザのテーブルの更新が完了しました。')

        select_sql = 'select * from users'
        for row in self.cursor.execute(select_sql):
            print(row)
        workbook.save(self.moneyBook)
    def Search(self,word1:str,word2:str)-> list:
        createTwitterUserTable='''create table if not exists TwitterExistsUser(gen int,realname TEXT,twittername TEXT,money int,remarks TEXT,authority TEXT,UNIQUE (realname,twittername)) '''
        self.cursor.execute(createTwitterUserTable)
        uReturnist=[]
        select_sql = '''select * from users where twittername is not null'''
        for row in self.cursor.execute(select_sql):
            uReturnist.append(row)
            #print(row)
        select_sql='''insert or ignore into TwitterExistsUser(gen,realname,twittername,money,remarks,authority) values (?,?,?,?,?,?)'''
        self.cursor.executemany(select_sql,uReturnist)
        select_sql = '''select * from TwitterExistsUser'''
        for row in self.cursor.execute(select_sql):
            pass
        returnList=[]
        #(word1,word2)
        #(at,本名 or twitterID or all)
        #(Lthan,money)
        #(Hthan,money)
        #(equals,money)
        #(get,root)
        if(word1=='at'):
            if(word2=='all'):
                at_all_sql='''select * from TwitterExistsUser'''
                for data in self.cursor.execute(at_all_sql):
                    returnList.append(data)
                return returnList
            else:#本名 or twitterID
                print('{0}をTwitterIDから検索中...'.format(word2))
                twitterID_sql='''select twittername from TwitterExistsUser'''
                for data in self.cursor.execute(twitterID_sql):
                    #なんかカッコとかついてるので取る
                    ddata0=str(data).replace('(','')
                    ddata1=ddata0.replace(')','')
                    ddata2=ddata1.replace('\'','')
                    ddata_final=ddata2.replace(',','')
                    #ここでddata_finalがアレ
                    if(ddata_final==word2):
                        print('twitterIDが一致しました。 そのユーザを取得します。')
                        userSearch_sql='''select * from TwitterExistsUser where twittername='{0}' '''.format(word2)
                        if (not self.cursor.execute(userSearch_sql)):
                            raise ValueError('kasu')
                        reacher=False
                        for i in self.cursor.execute(userSearch_sql):
                            if(reacher):
                                raise AssertionError('二つ以上の要素を持ってしまう致命的なエラー')
                            returnList.append(i)
                            reacher=True
                        return returnList
                print('TwitterIDは一致しませんでした。本名から検索します。')
                #
                print('{0}を本名から検索中...'.format(word2))
                realname_sql='''select realname from TwitterExistsUser'''
                for data in self.cursor.execute(realname_sql):
                    #なんかカッコとかついてるので取る
                    ddata0=str(data).replace('(','')
                    ddata1=ddata0.replace(')','')
                    ddata2=ddata1.replace('\'','')
                    ddata_final=ddata2.replace(',','')
                    #ここでddata_finalがアレ
                    if(ddata_final==word2):
                        print('本名が一致しました。 そのユーザを取得します。')
                        userSearch_sql='''select * from TwitterExistsUser where realname='{0}' '''.format(word2)
                        if (not self.cursor.execute(userSearch_sql)):
                            raise ValueError('kasu')
                        reacher=False
                        for i in self.cursor.execute(userSearch_sql):
                            if(reacher):
                                raise AssertionError('二つ以上の要素を持ってしまう致命的なエラー')
                            returnList.append(i)
                            reacher=True
                        return returnList
        if(word1=='Lthan' or word1=='Hthan'):
            if(word1=='Lthan'):
                comparison='<'
            else:
                comparison='>'
            if(type(word2) is str):
                raise ValueError('SQLインジェクション的な操作は禁止されています。\n引数を確認して下さい。')
            call_sql=f'''select * from TwitterExistsUser where money{comparison}{word2}'''
            print(call_sql)
            for i in self.cursor.execute(call_sql):
                returnList.append(i)
            return returnList
        if(word1=='get'):
            if(word2=='root'):
                select_sql='''select * from TwitterExistsUser where authority=='su' '''
                print(select_sql)
                for i in self.cursor.execute(select_sql):
                    returnList.append(i)
                return returnList
        raise  ValueError('値見つからない')
    def RegisterOrChanger(self,Rname:str,NewTwitterName:str,register=False)-> str:
        if(not register):
            try:
                listy=self.Search('at',Rname)
                for list in listy:
                    Rname=str(list[1])
            except:
                return 'あなたの名前は元々データベースに登録されていません。'
        wb=openpyxl.load_workbook(self.twitterBook)
        sheet=wb['Sheet1']
        for i in range(2,200):#const
            if(sheet.cell(row=i,column=1).value!=None):
                print(sheet.cell(row=i,column=1).value)
                if(sheet.cell(row=i,column=1).value==Rname):
                    if(register):
                        if(sheet.cell(row=i,column=2).value==None):
                            sheet.cell(row=i,column=2,value=NewTwitterName)
                            wb.save(self.twitterBook)
                            return '登録完了しました！'

                        else:
                            wb.save(self.twitterBook)
                            return 'すでに登録されています。infoコマンドをお使いください。'
                    else:
                        if(sheet.cell(row=i,column=2).value!=None):
                            sheet.cell(row=i,column=2,value=NewTwitterName)
                            wb.save(self.twitterBook)
                            return '変更完了しました！'
                        else:
                            wb.save(self.twitterBook)
                            return 'あなたのデータは登録されていません。registerコマンドを使用してデータベースに登録を行ってください。'
        wb.save(self.twitterBook)
        return 'あなたが誰か判別することが出来ませんでした。管理者に確認することをお勧め致します。'


